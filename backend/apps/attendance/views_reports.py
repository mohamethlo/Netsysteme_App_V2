# apps/attendance/views_reports.py
"""
Rapport mensuel de pointage.
À inclure dans apps/attendance/urls.py via :
    path("", include("apps.attendance.urls_reports")),
"""
import calendar
import datetime
import io
import logging

from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Attendance

logger = logging.getLogger(__name__)
User   = get_user_model()

JOURS_FR  = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
LATE_TIME = datetime.time(9, 15)


# ── Helpers ────────────────────────────────────────────────────────────────────
def _is_admin(user) -> bool:
    return user.has_business_permission("all") or (
        hasattr(user, "role") and user.role and
        user.role.name in ("Administrateur", "Dev_administration", "administration")
    )


def _full_name(user) -> str:
    fn = getattr(user, "prenom", "") or user.first_name or ""
    ln = getattr(user, "nom",    "") or user.last_name  or ""
    return f"{fn} {ln}".strip() or user.username


def working_days(start: datetime.date, end: datetime.date, role_name: str) -> int:
    """
    Jours ouvrables entre start et end (inclus).
    Dev_administration → Lun-Ven (5j)
    Autres             → Lun-Sam (6j)
    """
    exclude_saturday = (role_name == "Dev_administration")
    count = 0
    cur   = start
    while cur <= end:
        wd = cur.weekday()
        if exclude_saturday:
            if wd < 5:
                count += 1
        else:
            if wd < 6:
                count += 1
        cur += datetime.timedelta(days=1)
    return count


def report_end_date(year: int, month: int) -> datetime.date:
    today      = datetime.date.today()
    last_day   = datetime.date(year, month, calendar.monthrange(year, month)[1])
    return today if (year == today.year and month == today.month) else last_day


def build_employee_data(employee, first_day: datetime.date, end: datetime.date) -> dict:
    """Construit les stats d'un employé pour la période donnée."""
    role_name    = employee.role.name if getattr(employee, "role", None) else ""
    total_wd     = working_days(first_day, end, role_name)
    attendances  = Attendance.objects.filter(
        user=employee, date__gte=first_day, date__lte=end
    ).select_related("work_location").order_by("date")

    total_hours   = sum(float(a.total_hours or 0) for a in attendances)
    days_present  = sum(1 for a in attendances if a.check_in)
    days_late     = sum(1 for a in attendances if a.check_in and
                        timezone.localtime(a.check_in).time() > LATE_TIME)
    days_absent   = max(0, total_wd - days_present)
    just_lates    = sum(1 for a in attendances if a.check_in and
                        timezone.localtime(a.check_in).time() > LATE_TIME and
                        a.notes and a.notes.strip())
    unjust_lates  = days_late - just_lates
    rate          = round(days_present / total_wd * 100, 1) if total_wd else 0.0

    # Statut mérite
    if rate >= 95 and days_late <= 2:
        status = "excellent"
    elif rate >= 85 and days_late <= 5:
        status = "bon"
    elif rate >= 70:
        status = "moyen"
    else:
        status = "problematique"

    # Détails quotidiens
    details = []
    for a in attendances:
        ci = timezone.localtime(a.check_in)   if a.check_in  else None
        co = timezone.localtime(a.check_out)  if a.check_out else None
        details.append({
            "date":        a.date.strftime("%d/%m/%Y"),
            "day_name":    JOURS_FR[a.date.weekday()],
            "check_in":    ci.strftime("%H:%M") if ci else "—",
            "check_out":   co.strftime("%H:%M") if co else "—",
            "total_hours": f"{float(a.total_hours):.1f}h" if a.total_hours else "—",
            "location":    a.work_location.name if a.work_location else "—",
            "is_late":     bool(ci and ci.time() > LATE_TIME),
            "notes":       a.notes or "—",
            "status":      a.status or "",
        })

    return {
        "employee": {
            "id":        employee.id,
            "name":      _full_name(employee),
            "prenom":    getattr(employee, "prenom", "") or employee.first_name or "",
            "nom":       getattr(employee, "nom",    "") or employee.last_name  or "",
            "role":      role_name or "N/A",
            "email":     employee.email or "N/A",
            "telephone": getattr(employee, "telephone", "") or "N/A",
        },
        "total_hours":      round(total_hours, 1),
        "working_days":     total_wd,
        "days_present":     days_present,
        "days_absent":      days_absent,
        "days_late":        days_late,
        "justified_lates":  just_lates,
        "unjustified_lates":unjust_lates,
        "attendance_rate":  rate,
        "status":           status,
        "attendance_details": details,
    }


def sort_key(d: dict):
    return (-d["attendance_rate"], d["unjustified_lates"], d["days_absent"], -d["total_hours"])


# ── API : rapport mensuel ──────────────────────────────────────────────────────
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def monthly_report_api(request):
    if not _is_admin(request.user):
        return Response({"error": "Accès refusé."}, status=403)

    year  = int(request.query_params.get("year",  datetime.date.today().year))
    month = int(request.query_params.get("month", datetime.date.today().month))

    try:
        first_day = datetime.date(year, month, 1)
        end       = report_end_date(year, month)
        last_day  = datetime.date(year, month, calendar.monthrange(year, month)[1])
    except ValueError:
        return Response({"error": "Date invalide."}, status=400)

    employees   = User.objects.filter(is_active=True).select_related("role")
    report_data = sorted([build_employee_data(e, first_day, end) for e in employees], key=sort_key)

    total = len(report_data)
    return Response({
        "report_data":      report_data,
        "year":             year,
        "month":            month,
        "month_name":       calendar.month_name[month],
        "first_day":        first_day.isoformat(),
        "last_day":         last_day.isoformat(),
        "report_end":       end.isoformat(),
        "total_employees":  total,
        "avg_hours":        round(sum(r["total_hours"] for r in report_data) / total, 1) if total else 0,
        "total_absences":   sum(r["days_absent"] for r in report_data),
        "total_lates":      sum(r["days_late"]   for r in report_data),
        "avg_working_days": round(sum(r["working_days"] for r in report_data) / total, 1) if total else 0,
    })


# ── Export PDF global ──────────────────────────────────────────────────────────
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_monthly_pdf(request):
    if not _is_admin(request.user):
        return HttpResponse("Accès refusé", status=403)

    year  = int(request.query_params.get("year",  datetime.date.today().year))
    month = int(request.query_params.get("month", datetime.date.today().month))

    try:
        first_day = datetime.date(year, month, 1)
        end       = report_end_date(year, month)
    except ValueError:
        return HttpResponse("Date invalide", status=400)

    employees   = User.objects.filter(is_active=True).select_related("role")
    report_data = sorted([build_employee_data(e, first_day, end) for e in employees], key=sort_key)

    return _generate_global_pdf(report_data, year, month)


# ── Export Excel global ────────────────────────────────────────────────────────
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_monthly_excel(request):
    if not _is_admin(request.user):
        return HttpResponse("Accès refusé", status=403)

    year  = int(request.query_params.get("year",  datetime.date.today().year))
    month = int(request.query_params.get("month", datetime.date.today().month))

    try:
        first_day = datetime.date(year, month, 1)
        end       = report_end_date(year, month)
    except ValueError:
        return HttpResponse("Date invalide", status=400)

    employees   = User.objects.filter(is_active=True).select_related("role")
    report_data = sorted([build_employee_data(e, first_day, end) for e in employees], key=sort_key)

    return _generate_excel(report_data, year, month)


# ── Export PDF individuel ──────────────────────────────────────────────────────
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_employee_pdf(request, employee_id):
    if not _is_admin(request.user):
        return HttpResponse("Accès refusé", status=403)

    year  = int(request.query_params.get("year",  datetime.date.today().year))
    month = int(request.query_params.get("month", datetime.date.today().month))

    try:
        employee  = User.objects.select_related("role").get(pk=employee_id)
        first_day = datetime.date(year, month, 1)
        end       = report_end_date(year, month)
    except (User.DoesNotExist, ValueError):
        return HttpResponse("Introuvable ou date invalide", status=404)

    data = build_employee_data(employee, first_day, end)
    return _generate_employee_pdf(data, year, month)


# ── Générateurs PDF / Excel ────────────────────────────────────────────────────
def _generate_global_pdf(report_data: list, year: int, month: int) -> HttpResponse:
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1*cm)
    stys = getSampleStyleSheet()
    els  = []

    title_s = ParagraphStyle("T", parent=stys["Heading1"], fontSize=16,
                              textColor=colors.HexColor("#2c3e50"), alignment=TA_CENTER, spaceAfter=20)
    els.append(Paragraph(f"Rapport Mensuel — {calendar.month_name[month]} {year}", title_s))
    els.append(Spacer(1, 0.5*cm))

    rows = [["#", "Employé", "Rôle", "Heures", "Présent", "Absent", "Retards", "Just.", "Taux %"]]
    for i, r in enumerate(report_data, 1):
        rank = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else str(i)
        rows.append([rank, r["employee"]["name"], r["employee"]["role"],
                     f"{r['total_hours']}h", r["days_present"], r["days_absent"],
                     r["days_late"], r["justified_lates"], f"{r['attendance_rate']}%"])

    t = Table(rows, colWidths=[1.5*cm, 4.5*cm, 2.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm, 2*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3498db")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME",   (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0),  10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (0, 3), colors.HexColor("#d4edda")),
        ("GRID",       (0, 0), (-1, -1), 1, colors.black),
        ("FONTNAME",   (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 4), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    els.append(t)
    doc.build(els)
    buf.seek(0)

    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="rapport_mensuel_{month}_{year}.pdf"'
    return resp


def _generate_employee_pdf(data: dict, year: int, month: int) -> HttpResponse:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    rate   = data["attendance_rate"]
    status = data["status"]
    STATUS_COLOR = {
        "excellent":    colors.HexColor("#28a745"),
        "bon":          colors.HexColor("#17a2b8"),
        "moyen":        colors.HexColor("#ffc107"),
        "problematique":colors.HexColor("#dc3545"),
    }
    sc = STATUS_COLOR.get(status, colors.grey)

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                              leftMargin=2*cm, rightMargin=2*cm)
    stys = getSampleStyleSheet()
    els  = []

    t_s = ParagraphStyle("T", parent=stys["Heading1"], fontSize=18,
                         textColor=colors.HexColor("#2c3e50"), alignment=TA_CENTER, spaceAfter=10)
    s_s = ParagraphStyle("S", parent=stys["Normal"],  fontSize=12,
                         textColor=colors.HexColor("#34495e"), alignment=TA_CENTER, spaceAfter=20)

    els.append(Paragraph("RAPPORT DE POINTAGE MENSUEL", t_s))
    els.append(Paragraph(
        f"{data['employee']['name']}<br/>{calendar.month_name[month].capitalize()} {year}", s_s))
    els.append(Spacer(1, 0.5*cm))

    def section(title, bg):
        st = Table([[Paragraph(f"<b>{title}</b>", stys["Normal"])]], colWidths=[17*cm])
        st.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), bg),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0), 12),
            ("LEFTPADDING",(0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1),  8),
        ]))
        return st

    def info_table(rows):
        t = Table(rows, colWidths=[5*cm, 12*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
            ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
            ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
            ("FONTSIZE",   (0, 0), (-1, -1), 10),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
            ("LEFTPADDING",(0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",(0,0),(-1,-1),  6),
        ]))
        return t

    # Infos employé
    els.append(section("INFORMATIONS EMPLOYÉ", colors.HexColor("#3498db")))
    emp = data["employee"]
    els.append(info_table([
        ["Nom complet :", emp["name"]],
        ["Rôle :",        emp["role"]],
        ["Email :",       emp["email"]],
        ["Téléphone :",   emp["telephone"]],
    ]))
    els.append(Spacer(1, 0.5*cm))

    # Stats
    els.append(section("STATISTIQUES DU MOIS", colors.HexColor("#27ae60")))
    stat_t = Table([
        ["Heures totales :",  f"{data['total_hours']}h"],
        ["Jours présents :",  str(data["days_present"])],
        ["Jours absents :",   str(data["days_absent"])],
        ["Retards :",         f"{data['days_late']} ({data['justified_lates']} just., {data['unjustified_lates']} non just.)"],
        ["Taux de présence :", f"{rate}%"],
        ["Évaluation :",      status.upper()],
    ], colWidths=[5*cm, 12*cm])
    stat_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ecf0f1")),
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, -1), "LEFT"),
        ("FONTSIZE",   (0, 0), (-1, -1), 10),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("LEFTPADDING",(0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1),  6),
        ("BACKGROUND", (1, 5), (1, 5), sc),
        ("TEXTCOLOR",  (1, 5), (1, 5), colors.white),
        ("FONTNAME",   (1, 5), (1, 5), "Helvetica-Bold"),
    ]))
    els.append(stat_t)
    els.append(Spacer(1, 0.5*cm))

    # Pointages
    els.append(section("DÉTAIL DES POINTAGES QUOTIDIENS", colors.HexColor("#e67e22")))
    els.append(Spacer(1, 0.2*cm))
    att_rows = [["Date", "Jour", "Entrée", "Sortie", "Heures", "Lieu", "Statut"]]
    for d in data["attendance_details"]:
        s = "RETARD" if d["is_late"] else ("PRÉSENT" if d["check_in"] != "—" else "ABSENT")
        att_rows.append([d["date"], d["day_name"][:3], d["check_in"], d["check_out"],
                         d["total_hours"], d["location"], s])

    att_t = Table(att_rows, colWidths=[2*cm, 1.8*cm, 2*cm, 2*cm, 2*cm, 4.2*cm, 3*cm])
    ts = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1),  5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
    ]
    for idx, d in enumerate(data["attendance_details"], start=1):
        if d["is_late"]:
            ts += [("BACKGROUND",(6,idx),(6,idx), colors.HexColor("#fff3cd")),
                   ("TEXTCOLOR", (6,idx),(6,idx), colors.HexColor("#856404"))]
        elif d["check_in"] == "—":
            ts += [("BACKGROUND",(6,idx),(6,idx), colors.HexColor("#f8d7da")),
                   ("TEXTCOLOR", (6,idx),(6,idx), colors.HexColor("#721c24"))]
        else:
            ts += [("BACKGROUND",(6,idx),(6,idx), colors.HexColor("#d4edda")),
                   ("TEXTCOLOR", (6,idx),(6,idx), colors.HexColor("#155724"))]
    att_t.setStyle(TableStyle(ts))
    els.append(att_t)

    # Pied
    els.append(Spacer(1, 1*cm))
    foot_s = ParagraphStyle("F", parent=stys["Normal"], fontSize=8,
                             textColor=colors.grey, alignment=TA_CENTER)
    els.append(Paragraph(
        f"Document généré le {datetime.datetime.now().strftime('%d/%m/%Y à %H:%M')}<br/>"
        "Système de Gestion — Rapport automatisé", foot_s))

    doc.build(els)
    buf.seek(0)
    fname = f"rapport_{emp['prenom']}_{emp['nom']}_{month}_{year}.pdf".replace(" ", "_")
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


def _generate_excel(report_data: list, year: int, month: int) -> HttpResponse:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rapport {month}-{year}"

    h_fill   = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    h_font   = Font(bold=True, color="FFFFFF", size=12)
    top3_fill= PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    border   = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))

    ws.merge_cells("A1:I1")
    ws["A1"] = f"Rapport Mensuel — {calendar.month_name[month]} {year}"
    ws["A1"].font      = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["#", "Employé", "Rôle", "Heures", "Présent", "Absent", "Retards", "Justifiés", "Taux %"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = h_fill; c.font = h_font; c.border = border
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(report_data, 4):
        rank  = i - 3
        rdsp  = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉" if rank == 3 else str(rank)
        vals  = [rdsp, r["employee"]["name"], r["employee"]["role"],
                 r["total_hours"], r["days_present"], r["days_absent"],
                 r["days_late"], r["justified_lates"], r["attendance_rate"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = border
            if rank <= 3:
                c.fill = top3_fill

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="rapport_mensuel_{month}_{year}.xlsx"'
    return resp